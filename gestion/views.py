
from django import forms
from django.http import request
from django.shortcuts import redirect, render
from django.urls import reverse
import os
from datetime import date
from datetime import datetime  
from django.utils import timezone 
import pytz


from django.views.generic.edit import DeleteView, UpdateView
from openpyxl.styles.fills import FILL_NONE

from gestion.forms import activ_principalForm, sud_actividadForm , sud_actividadForm2 , Crea_Usuario , update_Usuario , update_contraseña_Usuario

from gestion.models import *
from gestion.Mixins import validarPermisosRequeridosMixin
from django.contrib.auth import authenticate, get_user_model

from django.views.generic import View, UpdateView, CreateView,  DetailView, TemplateView , ListView , DeleteView
from django.contrib.auth import authenticate , login , logout
from django.contrib.auth.decorators import login_required

from django.urls import  reverse_lazy


from  openpyxl import Workbook, workbook
from openpyxl.styles import  PatternFill, Border, Side, Alignment, Protection, Font, alignment
from openpyxl.drawing.image import Image
from django.http.response import HttpResponse, StreamingHttpResponse
from django.contrib import messages




from django.contrib.auth.mixins import PermissionRequiredMixin

from django.contrib.auth.mixins import LoginRequiredMixin





	  

def login_view(request):
	
	
	if request.method== 'POST':
		
		
		username = request.POST['email']
		password = request.POST['password']
	
		User = authenticate( username=username, password = password)
		
		
		if User :
			
			login(request , User)
		   
			
			return redirect('gestion:lista_actividades')
		else:
			return render(request , "corp/login.html" , {'error': 'Correo o Contraseña Invalido'})	
	return render(request , "corp/login.html")


@login_required
def logout_view(request):
	logout(request)
	return redirect('login')    






class crear_usuario( LoginRequiredMixin , validarPermisosRequeridosMixin , View):
	permission_required = 'gestion.add_user' 

   
	template_name = 'corp/signup.html'
	form_class = Crea_Usuario





	def get (self, request):

		return render(request,self.template_name , {'form':self.form_class})


	def post(self , request , *args , **kwargs):

		form  = self.form_class(request.POST) 
		

		if form.is_valid():
			

		 

			datos_limpios = form.cleaned_data

		  

		   

		   
			email = datos_limpios['email']
			first_name = datos_limpios['first_name']
			last_name = datos_limpios['last_name']
			password = datos_limpios['password']
			password_confirmation = datos_limpios['password_confirmation']
			cedula= datos_limpios['cedula']
			nombre_corporativo= datos_limpios['nombre_corporativo']
			estado= datos_limpios['estado']
			cargo = datos_limpios['cargo']
			

			
			for x in datos_limpios['groups']:

				
				
				if x.name == "nivel2" and  estado !="Aragua":
				   
					messages.error(request, " Usuario de nivel 2 solo debe ser del estado aragua") 
					return render(request , self.template_name , {'form':form})
					
					
				User = get_user_model()
				User = User.objects.create_user(username= nombre_corporativo, email=email , last_name =last_name , password = password , is_active = True ,
				cedula =cedula , first_name = first_name , estado =estado ,
				cargo = cargo)
				User.groups.add(x)



			"""User.set_password(password)
			User.is_active = True
			User.save()"""
			return redirect('gestion:listar_usiarios')

		return render(request , self.template_name , {'form':form})



	




class signup (View):

	def get (self, request):

		return render(request,'corp/signup.html')

	def post(self,request):

		username = request.POST['username']
		email = request.POST['username']

		
		nombre = request.POST['Nombre']
		last_name = request.POST['last_name']
		region = request.POST['estado']
		nivel = request.POST['nivel']
		

		
		password = request.POST['password']
		password_confirmation = request.POST['password_confirmation']

		context = {
			'fieldValues': request.POST
		}


		if password != password_confirmation:

			

			return render (request, 'corp/signup.html', {'error': 'Contraseñas no coinciden'})
		
		User = get_user_model()

		if User.objects.filter(email =email).exists():
			return render (request, 'corp/signup.html', {'error': 'Correo en uso'})

		if not User.objects.filter(email=email).exists():

		  User = User.objects.create_user(username=username, email=email)
		  User.set_password(password)
		  User.is_active = True
		  User.save()



		  profile= Profile( user=User, nombre=nombre , apellido=apellido , email=email, region=region, nivel=nivel)
		 
		  profile.save

		  return redirect('login')


		return render(request, 'corp/signup.html')






"""
	##############

			COMIENZO DESDE CERO

							###########################

					"""
  





class crear_actividad (LoginRequiredMixin , CreateView):
   
  
	model = activ_principal
	template_name = "corp/registro2.html"
	form_class =  activ_principalForm
	success_url = reverse_lazy('gestion:lista_actividades')




	def post(self , request , *args , **kwargs):
		form  = self.form_class(request.POST)
		queryset= User.objects.all()

		#asi se puede acceder a los grupos
		dato =0 
		for x in queryset:
			if x.groups.filter(user=self.request.user.id):
				dato = x.groups.filter(user=self.request.user.id)

		"""con el for de arriba guardo en la variable dato  un queryset o diccionario que contiene
		el  gupo al que pertenece el usuario y en 
		el siguiente for lo recorro para acceder a ese nombre y trabajarlo en la condicional de abajo"""

		nombre__de_grupo ="" 
		if dato !=0:
			for x in dato:
				nombre__de_grupo = x.name			

		if form.is_valid():

			estado_usuario = self.request.user.estado

			if nombre__de_grupo == "nivel1" and estado_usuario != form.cleaned_data['id_estado2']:
				messages.error(request, "Solo tiene permitido agregar actividades correspondientes al estado que usted pertenece") 
				return render(request , self.template_name , {'form':form})
			form.save()

			return redirect('gestion:lista_actividades')

		return render(request , self.template_name , {'form':form})


class crear_sud_actividad(LoginRequiredMixin , View):
   
	model =  sud_actividad
	second_model = activ_principal
	template_name = "corp/registro1.html"
	form_class =  sud_actividadForm2
	success_url = reverse_lazy('gestion:lista_actividades')


	#esta funcion me permite agarrar el valor de id de la actividad que viene en la url
	# con esto lo agarro self.kwargs['pk']
	def get_object(self , **kwargs):

		
		var = self.second_model.objects.get(id=self.kwargs['pk'])		

		return var


	"""
	esta funcion me permire reescribir el metodo post de mi clase y lo uso para guardar la sub actividad 
	lo hago asi por que quiero que el cliente no tenga necesidad de seleccionar a que actividad
	pertenece esta sub actividad 
	entonces con la funcion  get_object me traigo el objeto  que es la actividad para tener ese valor por defecto
	y asi el cliente no tendra que buscar pot ejemplo entre 1.000 actividades 
	para ver cual es la suya o cosas por el estilo"""

	def post(self , request , *args , **kwargs) :

		form = self.form_class(request.POST)

		if form.is_valid():
			
			#le quito el html al formulario que viene por el metodo post 
			formulario_limpio = form.cleaned_data


			#creo un objeto de mi modelo para usarlo para guardar 
			sub_actividad_guuardar = self.model()

			variable = self.get_object()

			#accedo a las propiedades de mi objeto
			sub_actividad_guuardar.num_actividad = formulario_limpio['num_actividad']
			sub_actividad_guuardar.nom_actividad= formulario_limpio['nom_actividad']
			sub_actividad_guuardar.fecha_inicio = formulario_limpio['fecha_inicio']
			sub_actividad_guuardar.fecha_fin = formulario_limpio['fecha_fin']
			sub_actividad_guuardar.fecha_fin_real = formulario_limpio['fecha_fin_real']

			sub_actividad_guuardar.estado = variable.id_estado2
			sub_actividad_guuardar.impacto = formulario_limpio['impacto']
			sub_actividad_guuardar.punto_critico = formulario_limpio['punto_critico']
			sub_actividad_guuardar.avance_programado = formulario_limpio['avance_programado']
			sub_actividad_guuardar.avance_ejecutado = formulario_limpio['avance_ejecutado']
			sub_actividad_guuardar.id_activ = self.get_object()

			sub_actividad_guuardar.save()

			#reverse contruye una url  lo uso ya que redirec no permite kwargs
			url = reverse('gestion:detalle_actividad' , kwargs={'pk':self.kwargs['pk']})
			
			return redirect(url)

		
		
			
		return render(request, 'corp/editar_subActividad.html' , {'form':form , 'object': self.get_object})
		
			

	def get_context_data(self , **kwargs):

		context= {}
	
		context['form']=self.form_class
		context['object']= self.get_object()
		
		return context



	def  get(self , request , *args , **kwargs):
		


		return render(request , self.template_name , self.get_context_data())        
			







class lista_actividades(LoginRequiredMixin , ListView):
	
	model = activ_principal
	template_name = "corp/listar_actividad.html"


	def get_context_data(self , **kwargs):
		
		context = super().get_context_data(**kwargs)
		
		variable = self.request.user.estado

		if self.request.user.is_staff or variable == "Aragua" :
			consulta = self.model.objects.all().order_by('-creado')
			
		else:
			consulta = self.model.objects.filter(id_estado2=variable).order_by('-creado')

		
		

		context['object_list'] = consulta
	   

		return context



class editar_actividad(LoginRequiredMixin , UpdateView):
	model= activ_principal
	form_class =  activ_principalForm
	template_name="corp/modal_editar_actividad.html"
	#fields=['num_actividades','nom_actividades','indicadores','costo','avance_1','alcance','region', 'id_estado2']  
	success_url = reverse_lazy('gestion:lista_actividades')
	

#validarPermisosRequeridosMixin es un validador de permisos creado por mi y esta en Mixins.py
class eliminar_actividad( LoginRequiredMixin , validarPermisosRequeridosMixin , DeleteView):
	permission_required = 'gestion.delete_activ_principal'
	model = activ_principal

	template_name="corp/eliminar_actividad.html"
	success_url = reverse_lazy('gestion:lista_actividades')



class detalle_actividad(LoginRequiredMixin , DetailView):
	second_model = sud_actividad
   
	template_name= 'corp/detalle_actvidad.html'
	pk_url_kwargs= 'pk'	
	queryset= activ_principal.objects.all()


	"""
	las primeras lineas me traen solo la actividad que selecciono al darle al boton
	 detalle en la lista de actividades
	y la siguiente funcion a continuacion lo que hago es traerme las actividades
	 a de esa actividad y mardarlas al templete para
	listarlas"""

	def get_context_data(self , **kwargs):

		context = super().get_context_data(**kwargs)
		
		datos = self.second_model.objects.filter(id_activ=self.kwargs['pk'])

		context['info']= datos

		return context


class editar_sub_actividad(LoginRequiredMixin , UpdateView):

	model= sud_actividad
	form_class = sud_actividadForm   
	template_name="corp/editar_subActividad.html"
	def post(self , request , *args , **kwargs):

		instancia = self.model.objects.get(id=self.kwargs['pk'])
		form = self.form_class(request.POST , instance=instancia)
		if form.is_valid():
			instancia = form.save(commit=False)
			# Podemos guardarla cuando queramos
			instancia.save()

			url = reverse('gestion:detalle_actividad' , kwargs={'pk':self.kwargs['pk2']})
			return redirect(url)

		return render(request , self.template_name , {'form':form})    
	

	

#validarPermisosRequeridosMixin es un validador de permisos creado por mi y esta en Mixins.py 
class eliminar_sub_Actividad(LoginRequiredMixin , validarPermisosRequeridosMixin , DeleteView):
	permission_required = 'gestion.delete_sud_actividad'
	model = sud_actividad
	template_name="corp/eliminar_subActividad.html"
	success_url = reverse_lazy('gestion:lista_actividades') 



	def post(self , request , *args , **kwargs):

		instancia = self.model.objects.get(id=self.kwargs['pk'])
		
		
		instancia.delete()
		   

		url = reverse('gestion:detalle_actividad' , kwargs={'pk':self.kwargs['pk2']})
		return redirect(url)





class listar_user( LoginRequiredMixin , validarPermisosRequeridosMixin , ListView):
	permission_required =   'gestion.view_user'

	model = User
	template_name = 'corp/lista_user.html'


class detalle_Usuario(LoginRequiredMixin ,  DetailView):

	template_name= 'corp/detalle_usuario.html'
	pk_url_kwargs= 'pk'	
	queryset= User.objects.all()

	#asi se puede acceder a los grupos
	"""dato =0 
		for x in self.queryset:
			if x.groups.filter(user=self.kwargs['pk']):
				dato = x.groups.filter(user=self.kwargs['pk'])

		print(dato)  """


	def get_context_data(self , **kwargs):


		context = super().get_context_data(**kwargs)

	 
		dato =0 
		for x in self.queryset:
			
			dato = x.password

		       
	   
	   #context['object2']= User.groups.filter(user_id=1).exists()
		
		return context

class editar_usuario(LoginRequiredMixin , validarPermisosRequeridosMixin , UpdateView):
	permission_required =   'gestion.change_user'   

	model= User
	form_class =  update_Usuario
	template_name ='corp/editar_usuario.html'



	def post(self , request , *args , **kwargs):

		instancia = self.model.objects.get(id=self.kwargs['pk'])
		form = self.form_class(request.POST , instance=instancia)
		if form.is_valid():
		   
			instancia.save()

			url = reverse('gestion:detalle_usiarios' , kwargs={'pk':self.kwargs['pk']})
			return redirect(url)

		return render(request , self.template_name , {'form':form})
	







class editar_contraseña_usuario( LoginRequiredMixin , validarPermisosRequeridosMixin , UpdateView):
	permission_required =   'gestion.change_user'   

	model= User
	form_class =  update_contraseña_Usuario
	template_name ='corp/editarcontraseña_usuario.html'
	



  
	def post(self , request , *args , **kwargs):

		instancia = self.model.objects.get(id=self.kwargs['pk'])
		form = self.form_class(request.POST , instance=instancia)
		if form.is_valid():
			
			
			instancia.set_password(form.cleaned_data.get('password'))

			

			instancia.save()

			url = reverse('gestion:detalle_usiarios' , kwargs={'pk':self.kwargs['pk']})
			return redirect(url)

		return render(request , self.template_name , {'form':form})




class eliminar_Usuario(LoginRequiredMixin , validarPermisosRequeridosMixin , DeleteView):
	permission_required =   'gestion.delete_user' 

   
	model = User

	template_name="corp/eliminar_Usuario.html"
	success_url = reverse_lazy('gestion:listar_usiarios')  




class genera_reporte ( LoginRequiredMixin , TemplateView):

	template_name="corp/genera_reporte.html"

class reporte_excel( LoginRequiredMixin ,  TemplateView):

	def get(self , request , *args , **kwargs):


		#variable utilizada como almacen de ese dato que describe el nombre de la variable 
		variable_calculo_porcentaje_total =0


		variable = self.request.user.estado

	

	
		#query = Order.objects.filter(ordered_date__range=[dato, dato2])
		if request.GET.get("customCheck1"):
			

			anio = int(request.GET.get("prueba2")) 
			primera_fecha = datetime(anio , 1 , 1 , tzinfo=pytz.UTC)
			segunda_fecha = datetime(anio , 4 , 30 , tzinfo=pytz.UTC)

			if self.request.user.is_staff or variable == "Aragua" :
				
				#imprimo todas las actividades 
				query = activ_principal.objects.filter(creado__range=[primera_fecha, segunda_fecha])
			else:
				#filtro las actividades dependiendo del estado 
				query = activ_principal.objects.filter(creado__range=[primera_fecha, segunda_fecha] , id_estado2=variable)

			
			
		else:
			if 	request.GET.get("customCheck2"):
				
				anio = int(request.GET.get("prueba2")) 
				primera_fecha = datetime(anio , 5 , 1 , tzinfo=pytz.UTC)
				segunda_fecha = datetime(anio , 8 , 31 , tzinfo=pytz.UTC)

				if self.request.user.is_staff or variable == "Aragua" :
				
					#imprimo todas las actividades 
					query = activ_principal.objects.filter(creado__range=[primera_fecha, segunda_fecha])
				else:
					#filtro las actividades dependiendo del estado 
					query = activ_principal.objects.filter(creado__range=[primera_fecha, segunda_fecha] , id_estado2=variable)




				
				
				
				
			else:
				if request.GET.get("customCheck3"):
					
					anio = int(request.GET.get("prueba2")) 
					primera_fecha = datetime(anio , 9 , 1 , tzinfo=pytz.UTC)
					segunda_fecha = datetime(anio , 12 , 31 , tzinfo=pytz.UTC)

					if self.request.user.is_staff or variable == "Aragua" :
				
						#imprimo todas las actividades 
						query = activ_principal.objects.filter(creado__range=[primera_fecha, segunda_fecha])
					else:

						#filtro las actividades dependiendo del estado 
						query = activ_principal.objects.filter(creado__range=[primera_fecha, segunda_fecha] , id_estado2=variable)


					
				else:
					if request.GET.get("customCheck4"):
						
						anio = int(request.GET.get("prueba2")) 
						primera_fecha = datetime(anio , 1 , 1 , tzinfo=pytz.UTC)
						segunda_fecha = datetime(anio , 12 , 31 , tzinfo=pytz.UTC)
						
						if self.request.user.is_staff or variable == "Aragua" :
				
							#imprimo todas las actividades 
							query = activ_principal.objects.filter(creado__range=[primera_fecha, segunda_fecha])
						else:
							#filtro las actividades dependiendo del estado 
							query = activ_principal.objects.filter(creado__range=[primera_fecha, segunda_fecha] , id_estado2=variable)


						



		
		
		currentDir= os.path.abspath(os.path.dirname(__file__)) 
		filepath=os.path.join(currentDir,"panaderiaG.jpg")
		#if not dato and not dato2:
		


	

		wb= Workbook()
		ws = wb.active
		#bandera = True
		
		controlador = 4


		img = Image(filepath)
		ws.add_image(img, 'A1')
		wb.save('panaderiaG.jpg.xlsx')
		
		
		#print(sheet.colum)
		ws['B1'].alignment = Alignment(horizontal = "center",vertical = "center")
		ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )

		ws['B1'].fill = PatternFill(start_color = 'f70202', end_color = 'f70202', fill_type = "solid")
		ws['B1'].font = Font(name = 'Calibri', size = 20, bold = True)

		ws['B1'] =  'CORPOELEC Plan De Accion ATIT'


		ws.merge_cells('B1:E1')

		ws.row_dimensions[1].height = 50
		ws.row_dimensions[3].height = 30

		ws.column_dimensions['B'].width = 30
		ws.column_dimensions['C'].width = 30
		ws.column_dimensions['D'].width = 30
		ws.column_dimensions['E'].width = 30
		ws.column_dimensions['F'].width = 30
		ws.column_dimensions['G'].width = 30
		ws.column_dimensions['H'].width = 30
		ws.column_dimensions['I'].width = 30
		ws.column_dimensions['J'].width = 30

		ws.column_dimensions['K'].width = 30
		ws.column_dimensions['L'].width = 30
		
		
		

		ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
		ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws['B3'].fill = PatternFill(start_color = '0550e8', end_color = '0550e8', fill_type = "solid")
		ws['B3'].font = Font(name = 'Calibro', size = 10, bold = True)
		ws['B3'] = 'num actividades'

		ws['C3'].alignment = Alignment(horizontal = "center", vertical = "center")
		ws['C3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws['C3'].fill = PatternFill(start_color = '0550e8', end_color = '0550e8', fill_type = "solid")
		ws['C3'].font = Font(name = 'Calibro', size = 10, bold = True)
		ws['C3'] = 'nom actividades'



		ws['D3'].alignment = Alignment(horizontal = "center", vertical = "center")
		ws['D3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws['D3'].fill = PatternFill(start_color = '0550e8', end_color = '0550e8', fill_type = "solid")
		ws['D3'].font = Font(name = 'Calibro', size = 10, bold = True)
		ws['D3'] = 'Indicadores'
		ws['E3'].alignment = Alignment(horizontal = "center", vertical = "center")
		ws['E3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws['E3'].fill = PatternFill(start_color = '0550e8', end_color = '0550e8', fill_type = "solid")
		ws['E3'].font = Font(name = 'Calibro', size = 10, bold = True)
		ws['E3'] = 'Region'

		ws['F3'].alignment = Alignment(horizontal = "center", vertical = "center")
		ws['F3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws['F3'].fill = PatternFill(start_color = '0550e8', end_color = '0550e8', fill_type = "solid")
		ws['F3'].font = Font(name = 'Calibro', size = 10, bold = True)
		ws['F3'] = 'Estado'

		ws['G3'].alignment = Alignment(horizontal = "center", vertical = "center")
		ws['G3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws['G3'].fill = PatternFill(start_color = '0550e8', end_color = '0550e8', fill_type = "solid")
		ws['G3'].font = Font(name = 'Calibro', size = 10, bold = True)
		ws['G3'] = 'Fecha Inicio'

		ws['H3'].alignment = Alignment(horizontal = "center", vertical = "center")
		ws['H3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws['H3'].fill = PatternFill(start_color = '0550e8', end_color = '0550e8', fill_type = "solid")
		ws['H3'].font = Font(name = 'Calibro', size = 10, bold = True)
		ws['H3'] = 'Fecha fin'


		ws['I3'].alignment = Alignment(horizontal = "center", vertical = "center")
		ws['I3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws['I3'].fill = PatternFill(start_color = '0550e8', end_color = '0550e8', fill_type = "solid")
		ws['I3'].font = Font(name = 'Calibro', size = 10, bold = True)
		ws['I3'] = 'Avance fisico programado'



		ws['J3'].alignment = Alignment(horizontal = "center", vertical = "center")
		ws['J3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws['J3'].fill = PatternFill(start_color = '0550e8', end_color = '0550e8', fill_type = "solid")
		ws['J3'].font = Font(name = 'Calibro', size = 10, bold = True)
		ws['J3'] = 'Avance fisico ejecutado'



		ws['K3'].alignment = Alignment(horizontal = "center", vertical = "center")
		ws['K3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
								   top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws['K3'].fill = PatternFill(start_color = '0550e8', end_color = '0550e8', fill_type = "solid")
		ws['K3'].font = Font(name = 'Calibro', size = 10, bold = True)
		ws['K3'] = 'Impacto'




		ws['L3'].alignment = Alignment(horizontal = "center", vertical = "center")
		ws['L3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
								   top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws['L3'].fill = PatternFill(start_color = '0550e8', end_color = '0550e8', fill_type = "solid")
		ws['L3'].font = Font(name = 'Calibro', size = 10, bold = True)
		ws['L3'] = 'Punto criticio'



	   



		

		cont= 1
		for q in query:
			ws.row_dimensions[controlador].height = 30
			
			

			
		   
		   
			#Pintamos los datos en el reporte
			
			ws.cell(row = controlador , column = 2).alignment = Alignment(horizontal = "center")
			ws.cell(row = controlador , column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
			ws.cell(row = controlador, column = 2).font = Font(name = 'Calibri', size = 8)
			ws.cell(row = controlador, column = 2).fill = PatternFill(start_color = '949aa6', end_color = '949aa6', fill_type = "solid")
			ws.cell(row = controlador, column = 2).value = q.num_actividades

			ws.cell(row = controlador, column = 3).alignment = Alignment(horizontal = "center")
			ws.cell(row = controlador, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
																	top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
			ws.cell(row = controlador, column = 3).font = Font(name = 'Calibri', size = 8)
			ws.cell(row = controlador, column = 3).fill = PatternFill(start_color = '949aa6', end_color = '949aa6', fill_type = "solid")

			ws.cell(row = controlador, column = 3).value = q.nom_actividades

			ws.cell(row = controlador, column = 4).alignment = Alignment(horizontal = "center")
			ws.cell(row = controlador, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
			ws.cell(row = controlador, column = 4).font = Font(name = 'Calibri', size = 8)
			ws.cell(row = controlador, column = 4).fill = PatternFill(start_color = '949aa6', end_color = '949aa6', fill_type = "solid")

			ws.cell(row = controlador, column = 4).value = q.indicadores


			ws.cell(row = controlador, column = 5).alignment = Alignment(horizontal = "center")
			ws.cell(row = controlador, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
			ws.cell(row = controlador, column = 5).font = Font(name = 'Calibri', size = 8)
			ws.cell(row = controlador, column = 5).fill = PatternFill(start_color = '949aa6', end_color = '949aa6', fill_type = "solid")

			ws.cell(row = controlador, column = 5).value = q.region




			ws.cell(row = controlador, column = 6).alignment = Alignment(horizontal = "center")
			ws.cell(row = controlador, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
			ws.cell(row = controlador, column = 6).font = Font(name = 'Calibri', size = 8)
			ws.cell(row = controlador, column = 6).fill = PatternFill(start_color = '949aa6', end_color = '949aa6', fill_type = "solid")

			ws.cell(row = controlador, column = 6).value = q.id_estado2



			ws.cell(row = controlador, column = 7).alignment = Alignment(horizontal = "center")
			ws.cell(row = controlador, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
			ws.cell(row = controlador, column = 7).font = Font(name = 'Calibri', size = 8)
			ws.cell(row = controlador, column = 7).fill = PatternFill(start_color = '949aa6', end_color = '949aa6', fill_type = "solid")

			ws.cell(row = controlador, column = 7).value = ""


			ws.cell(row = controlador, column = 8).alignment = Alignment(horizontal = "center")
			ws.cell(row = controlador, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
			ws.cell(row = controlador, column = 8).font = Font(name = 'Calibri', size = 8)
			ws.cell(row = controlador, column = 8).fill = PatternFill(start_color = '949aa6', end_color = '949aa6', fill_type = "solid")

			ws.cell(row = controlador, column = 8).value = ""

			#calculo del promedio de avance programado 
			try:
				query4 = sud_actividad.objects.filter(id_activ = q.id)

				suma=0
				for datos in query4:
					suma+=datos.avance_programado

		
				calculo_avance_programado =	suma/len(query4)
				con_dos_decimales_calculo_avance_programado = round(calculo_avance_programado, 2)

			except:		
				con_dos_decimales_calculo_avance_programado = ""



			ws.cell(row = controlador, column = 9).alignment = Alignment(horizontal = "center")
			ws.cell(row = controlador, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
			ws.cell(row = controlador, column = 9).font = Font(name = 'Calibri', size = 8)
			ws.cell(row = controlador, column = 9).fill = PatternFill(start_color = '949aa6', end_color = '949aa6', fill_type = "solid")

			ws.cell(row = controlador, column = 9).value = con_dos_decimales_calculo_avance_programado

			
			
			
			
			
			#calculo del promedio de avance ejecutado
			try:

				query3 = sud_actividad.objects.filter(id_activ = q.id)

				suma=0
				for datos in query3:
					suma+=datos.avance_ejecutado

		
				calculo_avance_ejecutado =	suma/len(query3)
				con_dos_decimales_calculo_avance_ejecutado = round(calculo_avance_ejecutado, 2)

				variable_calculo_porcentaje_total += con_dos_decimales_calculo_avance_ejecutado 	
			except:
				con_dos_decimales_calculo_avance_programado = ""






			ws.cell(row = controlador, column = 10).alignment = Alignment(horizontal = "center")
			ws.cell(row = controlador, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
			ws.cell(row = controlador, column = 10).font = Font(name = 'Calibri', size = 8)
			ws.cell(row = controlador, column = 10).fill = PatternFill(start_color = '949aa6', end_color = '949aa6', fill_type = "solid")

			ws.cell(row = controlador, column = 10).value = con_dos_decimales_calculo_avance_ejecutado



			ws.cell(row = controlador, column = 11).alignment = Alignment(horizontal = "center")
			ws.cell(row = controlador, column = 11).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
			ws.cell(row = controlador, column = 11).font = Font(name = 'Calibri', size = 8)
			ws.cell(row = controlador, column = 11).fill = PatternFill(start_color = '949aa6', end_color = '949aa6', fill_type = "solid")

			ws.cell(row = controlador, column = 11).value = ""



			ws.cell(row = controlador, column = 12).alignment = Alignment(horizontal = "center")
			ws.cell(row = controlador, column = 12).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
			ws.cell(row = controlador, column = 12).font = Font(name = 'Calibri', size = 8)
			ws.cell(row = controlador, column = 12).fill = PatternFill(start_color = '949aa6', end_color = '949aa6', fill_type = "solid")

			ws.cell(row = controlador, column = 12).value = ""

			


			query2 = sud_actividad.objects.filter(id_activ = q.id)
			
			for datos in query2:

				controlador+=1
				ws.row_dimensions[controlador].height = 50
				
				ws.cell(row = controlador , column = 2).alignment = Alignment(horizontal = "center")
				ws.cell(row = controlador , column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
										top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
				ws.cell(row = controlador, column = 2).font = Font(name = 'Calibri', size = 8)
				ws.cell(row = controlador, column = 2).value = datos.num_actividad

				ws.cell(row = controlador, column = 3).alignment = Alignment(horizontal = "center")
				ws.cell(row = controlador, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
																		top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
				ws.cell(row = controlador, column = 3).font = Font(name = 'Calibri', size = 8)
				ws.cell(row = controlador, column = 3).value = datos.nom_actividad

				ws.cell(row = controlador, column = 4).alignment = Alignment(horizontal = "center")
				ws.cell(row = controlador, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
										top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
				ws.cell(row = controlador, column = 4).font = Font(name = 'Calibri', size = 8)
				ws.cell(row = controlador, column = 4).value = ""


				ws.cell(row = controlador, column = 5).alignment = Alignment(horizontal = "center")
				ws.cell(row = controlador, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
										top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
				ws.cell(row = controlador, column = 5).font = Font(name = 'Calibri', size = 8)
				ws.cell(row = controlador, column = 5).value = "CENTRAL"




				ws.cell(row = controlador, column = 6).alignment = Alignment(horizontal = "center")
				ws.cell(row = controlador, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
										top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
				ws.cell(row = controlador, column = 6).font = Font(name = 'Calibri', size = 8)
				ws.cell(row = controlador, column = 6).value = datos.estado


				ws.cell(row = controlador, column = 7).alignment = Alignment(horizontal = "center")
				ws.cell(row = controlador, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
										top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
				ws.cell(row = controlador, column = 7).font = Font(name = 'Calibri', size = 8)
				ws.cell(row = controlador, column = 7).value = datos.fecha_inicio
				

				ws.cell(row = controlador, column = 8).alignment = Alignment(horizontal = "center")
				ws.cell(row = controlador, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
										top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
				ws.cell(row = controlador, column = 8).font = Font(name = 'Calibri', size = 8)
				ws.cell(row = controlador, column = 8).value = datos.fecha_fin

				ws.cell(row = controlador, column = 9).alignment = Alignment(horizontal = "center")
				ws.cell(row = controlador, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
										top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
				ws.cell(row = controlador, column = 9).font = Font(name = 'Calibri', size = 8)
				ws.cell(row = controlador, column = 9).value = datos.avance_programado

				ws.cell(row = controlador, column = 10).alignment = Alignment(horizontal = "center")
				ws.cell(row = controlador, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
										top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
				ws.cell(row = controlador, column = 10).font = Font(name = 'Calibri', size = 8)
				ws.cell(row = controlador, column = 10).value = datos.avance_ejecutado


				ws.cell(row = controlador, column = 11).alignment = Alignment(vertical = "distributed" )
				ws.cell(row = controlador, column = 11).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
										top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
				ws.cell(row = controlador, column = 11).font = Font(name = 'Calibri', size = 8)
				ws.cell(row = controlador, column = 11).value = datos.impacto

				ws.cell(row = controlador, column = 12).alignment = Alignment(horizontal = "center")
				ws.cell(row = controlador, column = 12).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
										top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
				ws.cell(row = controlador, column = 12).font = Font(name = 'Calibri', size = 8)
				ws.cell(row = controlador, column = 12).value = datos.punto_critico

				
				

			controlador +=1


		controlador+=1

		ws.cell(row = controlador, column = 9).alignment = Alignment(horizontal = "center")
		ws.cell(row = controlador, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws.cell(row = controlador, column = 9).font = Font(name = 'Calibri', size = 11)
		ws.cell(row = controlador, column = 9).fill = PatternFill(start_color = 'ed0909', end_color = 'ed0909', fill_type = "solid")

		ws.cell(row = controlador, column = 9).value = "Porcentaje de cumplimiento total:"

		""" DE AQUI SALE EL CALCULO FINAOL DEL PORCENTAJE DE TODAS LAS ACTIVIDADES 
		MOSTRADAS EN EL REPORTE"""



		resultado_final = variable_calculo_porcentaje_total/len(query)
		redondeo =  round(resultado_final, 2)
		

		ws.cell(row = controlador, column = 10).alignment = Alignment(horizontal = "center")
		ws.cell(row = controlador, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
									top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws.cell(row = controlador, column = 10).font = Font(name = 'Calibri', size = 10)
		ws.cell(row = controlador, column = 10).fill = PatternFill(start_color = 'ed0909', end_color = 'ed0909', fill_type = "solid")

		ws.cell(row = controlador, column = 10).value = redondeo



		#establecer nombre de archivo

		nombre_archivo ="PlanDeAccionATIT.xlsx"

		response = HttpResponse(content_type= 'application/ms-excel')
		contenido = 'attachment; filename = {0}'.format(nombre_archivo)
		response['Content-Disposition'] = contenido
		wb.save(response)
		return response	



class muestra_grafica( LoginRequiredMixin , View):
	model= activ_principal
	template_name="corp/muestra_grafica.html"



	def get(self , request , *args , **kwargs):


		primer_anio = request.GET.get('prueba1')
		segundo_anio = request.GET.get('prueba2')
		print( primer_anio ,  segundo_anio )

		
		query = self.model.objects.filter(creado__year = primer_anio)

		lista_de_porcentaje_por_anio= []
		try:
			for q in query:


				query3 = sud_actividad.objects.filter(id_activ = q.id)
				suma=0
				for datos in query3:
					suma+=datos.avance_ejecutado

			
				calculo_avance_ejecutado =	suma/len(query3)
				con_dos_decimales_calculo_avance_ejecutado = round(calculo_avance_ejecutado, 2)
				lista_de_porcentaje_por_anio.append(con_dos_decimales_calculo_avance_ejecutado)
		except:
			pass

		

		suma = 0
		if len(lista_de_porcentaje_por_anio)!=0:
			for datos in lista_de_porcentaje_por_anio:
				suma+=datos

			resultado_para_grafica =  suma/len(lista_de_porcentaje_por_anio)
			calculo_avance_ejecutado = round(resultado_para_grafica, 2)
			print(lista_de_porcentaje_por_anio)

			anio_2022 = {'anio_2022':calculo_avance_ejecutado}
			anio_2022['nombre1'] = primer_anio
		else:
				anio_2022 = {'anio_2022':0}
				anio_2022['nombre1'] = primer_anio






	#-------------------------------------------------------
		lista_de_porcentaje_por_anio2=[]
		query2 = self.model.objects.filter(creado__year = segundo_anio)
		
		try:
			for q in query2:


				query4 = sud_actividad.objects.filter(id_activ = q.id)
				suma2=0
				
				for datos in query4:
					suma2+=datos.avance_ejecutado

					
				calculo_avance_ejecutado =	suma2/len(query4)
				con_dos_decimales_calculo_avance_ejecutado = round(calculo_avance_ejecutado, 2)
				
				lista_de_porcentaje_por_anio2.append(con_dos_decimales_calculo_avance_ejecutado)
		except:
			pass

		

		if len(lista_de_porcentaje_por_anio2)!=0:
			suma2 = 0
			for datos in lista_de_porcentaje_por_anio2:
				suma2+=datos

			resultado_para_grafica =  suma2/len(lista_de_porcentaje_por_anio2)
			calculo_avance_ejecutado = round(resultado_para_grafica, 2)
			

			anio_2022["segundo_anio"] = calculo_avance_ejecutado
			anio_2022['nombre2'] = segundo_anio
		else:
			anio_2022["segundo_anio"] = 0
			anio_2022['nombre2'] = segundo_anio

	






		
		return render(request , self.template_name , {'anio_2022':anio_2022})



class formulario_estadistica( LoginRequiredMixin , TemplateView):

	template_name = "corp/formulario_estadisticas_consulta.html"





	 


  



	


	


		







		

	



